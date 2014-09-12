#-*- coding: utf-8 -*-
import pgdb
import openpyxl
import time
from openpyxl.reader.excel import load_workbook  
# import settings

import sys
default_encoding = 'utf-8'
if sys.getdefaultencoding() != default_encoding:
    reload(sys)
    sys.setdefaultencoding(default_encoding) 

# host = "42.121.18.61:5432"
# user = "soulagou_read"
# password = "123456789"
host = "192.168.19.245:5433"
user = "postgres"
password = "sale114"

db_outlet = "soulagou_outlet_new"
conn_outlet = pgdb.connect(database = db_outlet, host = host, user = user, password = password)
cursor_outlet = conn_outlet.cursor()
gender_table = [u'',u'男', u'女']
operators_table = [u'',u'移动', u'电信', u'联通']
def getTrial() :
	sql = "select t.id,t.name, t.gender,t.mobile,t.operators, t.create_time from trial_user_4g t"

	cursor_outlet.execute(sql)
	rs = cursor_outlet.fetchall()

	data = list()
	for r in rs:
		record = dict()
		record['ID'] = r[0]
		record['name'] = r[1]
		record['gender'] = gender_table[r[2]]
		record['mobile'] = r[3]
		record['operators'] = operators_table[r[4]]
		record['create_time'] = r[5]
		data.append(record)

	return data

def write_xls(data):
    
	wb = openpyxl.workbook.Workbook()
	s = wb.create_sheet(0)
	s.title = u'4G体验报名第一批'

	s.cell(row=1, column=1).value = u'序号'
	s.cell(row=1, column=2).value = u'姓名'
	s.cell(row=1, column=3).value = u'性别'
	s.cell(row=1, column=4).value = u'手机号码'
	s.cell(row=1, column=5).value = u'选择体验运营商'
	s.cell(row=1, column=6).value = u'报名时间'
	time_format = '%Y-%m-%d %X'
	r = 2;
	for i in data:
		try:
			s.cell(row=r, column=1).value = i['ID']
			s.cell(row=r, column=2).value = i['name'] 
			s.cell(row=r, column=3).value = i['gender'] 
			s.cell(row=r, column=4).value = i['mobile']
			s.cell(row=r, column=5).value = i['operators'] 
			s.cell(row=r, column=6).value = i['create_time'] 
			print i
		except Exception, e:
			print "PROCESS %s ERROR %s" % (i['ID'], e)
		r += 1
	xls_name = '%s.xls' % u'4G体验报名第一批名单'
	wb.save('%s/%s' % ('/home/lisong', xls_name))

def read_xls():
	xls_name = '%s.xls' % u'4G体验报名第一批名单'
	wb = load_workbook(filename = r'%s/%s' % ('/home/lisong', xls_name)) 
	sheetnames = wb.get_sheet_names() 
	ws = wb.get_sheet_by_name(sheetnames[1])  

	wbw = openpyxl.workbook.Workbook()
	s = wbw.create_sheet(0)
	s.title = u'4G体验报名第一批公布名单'

	records = ws.get_highest_row()  
	for r in range(2, records):
		try:
			user = dict()
			user['ID'] = ws.cell(row=r, column=1).value
			user['name']  = ws.cell(row=r, column=2).value
			user['gender']  = gender_table.index(ws.cell(row=r, column=3).value)
			user['mobile']  = ws.cell(row=r, column=4).value
			user['operators']  = operators_table.index(ws.cell(row=r, column=5).value)
			user['create_time']  = ws.cell(row=r, column=6).value
			write_db(user)
			print user
		except Exception, e:
			print "PROCESS ERROR %s" % e
	conn_outlet.commit()
	conn_outlet.close()
	# xls_name = '%s.xlsx' % u'4G体验报名第一批公布名单'
	# wbw.save('%s/%s' % ('/home/lisong', xls_name))
def write_db(user):
	sql = 'INSERT INTO trial_user_show (name, gender, mobile, operators, create_time) VALUES (%s, %s, %s, %s, %s)'
	cursor_outlet.execute(sql, [user['name'], user['gender'], user['mobile'], user['operators'], user['create_time']])
        
if __name__ == '__main__':
	# write_xls(getTrial())
	read_xls()
	print '--------------'
	# print gender_table.index(u'男')